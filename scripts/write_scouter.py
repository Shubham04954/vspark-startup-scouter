"""
write_scouter.py — Append rows into the New_Startup_Scouter sheet of
Ventures_Tracker.xlsx with correct column layout, per-column formatting,
emoji-based colour fills, and legend re-merge.

Sets explicit per-column formatting matching the real tracker:
  - Col B (#): white bold font on dark navy fill
  - Col C (Name): navy bold font on white fill
  - Cols D, G-L (data): navy font on white fill
  - Cols E, F (URLs): blue underlined font on white fill
  - Cols M-T (BU/Verdict): emoji-driven colour fill, dark font
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Fill colours
NAVY = PatternFill(start_color='FF0D2137', end_color='FF0D2137', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
GREEN = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')
AMBER = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
RED   = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Fonts
WHITE_FONT_BOLD = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
NAVY_FONT = Font(name='Calibri', size=9, color='FF0D2137')
NAVY_FONT_BOLD = Font(name='Calibri', size=9, bold=True, color='FF0D2137')
LINK_FONT = Font(name='Calibri', size=9, color='FF0000FF', underline='single')
DARK_FONT = Font(name='Calibri', size=9, color='FF0D2137')

# Alignment
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)
LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Border (thin all sides, light grey)
THIN = Side(style='thin', color='FFD9D9D9')
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def detect_variant(ws):
    """Read row 4 to find which column has BALCO header."""
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(row=4, column=c).value or '').strip().upper()
        first_line = header.split('\n')[0].strip()
        if first_line == 'BALCO':
            balco = c
            return {
                '#': balco - 11,
                'name': balco - 10,
                'indian': balco - 9,
                'website': balco - 8,
                'linkedin': balco - 7,
                'founded': balco - 6,
                'location': balco - 5,
                'funding': balco - 4,
                'founders': balco - 3,
                'team_size': balco - 2,
                'capabilities': balco - 1,
                'BALCO': balco,
                'VAL-L': balco + 1,
                'VAL-J': balco + 2,
                'ESL': balco + 3,
                'HZL': balco + 4,
                'Cairn': balco + 5,
                'IOB': balco + 6,
                'verdict': balco + 7,
            }
    raise ValueError("Could not find BALCO header in row 4")


def apply_emoji_fill(cell):
    """Set fill colour based on first emoji in cell value."""
    v = str(cell.value or '')[:3]
    if v.startswith('✅'):
        cell.fill = GREEN
    elif v.startswith('⚠️'):
        cell.fill = AMBER
    elif v.startswith('❌') or v.startswith('🚩') or v.startswith('🔍'):
        cell.fill = RED


def format_data_row(ws, r, cols):
    """Apply explicit per-column formatting to a data row."""
    # B = serial number — white bold on navy
    cell = ws.cell(row=r, column=cols['#'])
    cell.font = WHITE_FONT_BOLD
    cell.fill = NAVY
    cell.alignment = CENTER
    cell.border = ALL_BORDERS

    # C = company name — navy bold on white
    cell = ws.cell(row=r, column=cols['name'])
    cell.font = NAVY_FONT_BOLD
    cell.fill = WHITE_FILL
    cell.alignment = LEFT_CENTER
    cell.border = ALL_BORDERS

    # D = Indian? — navy on white, centred
    cell = ws.cell(row=r, column=cols['indian'])
    cell.font = NAVY_FONT
    cell.fill = WHITE_FILL
    cell.alignment = CENTER
    cell.border = ALL_BORDERS

    # E, F = URLs — blue link font on white
    for col_key in ['website', 'linkedin']:
        cell = ws.cell(row=r, column=cols[col_key])
        cell.font = LINK_FONT
        cell.fill = WHITE_FILL
        cell.alignment = LEFT_TOP
        cell.border = ALL_BORDERS

    # G-L = text data — navy on white
    for col_key in ['founded', 'location', 'funding', 'founders', 'team_size', 'capabilities']:
        cell = ws.cell(row=r, column=cols[col_key])
        cell.font = NAVY_FONT
        cell.fill = WHITE_FILL
        cell.alignment = LEFT_TOP
        cell.border = ALL_BORDERS

    # M-T = BU + Verdict — dark font, fill applied per-emoji separately
    for col_key in ['BALCO', 'VAL-L', 'VAL-J', 'ESL', 'HZL', 'Cairn', 'IOB', 'verdict']:
        cell = ws.cell(row=r, column=cols[col_key])
        cell.font = DARK_FONT
        cell.alignment = LEFT_TOP
        cell.border = ALL_BORDERS


def find_legend_row(ws, cols):
    """Find the row containing 'LEGEND' text."""
    serial_col = cols['#']
    for r in range(5, ws.max_row + 2):
        v = str(ws.cell(row=r, column=serial_col).value or '')
        if 'LEGEND' in v.upper():
            return r
    return None


def find_last_data_row(ws, cols):
    """Find the last row containing actual startup data."""
    serial_col = cols['#']
    last = 4
    for r in range(5, ws.max_row + 2):
        v = ws.cell(row=r, column=serial_col).value
        if v is None:
            continue
        if isinstance(v, (int, float)):
            last = r
        elif isinstance(v, str) and 'LEGEND' not in v.upper():
            name = str(ws.cell(row=r, column=cols['name']).value or '')
            if name and 'LEGEND' not in name.upper():
                last = r
    return last


def append_to_scouter(workbook_path, rows, sheet_name='New_Startup_Scouter'):
    """Append a list of row-dicts into the New_Startup_Scouter sheet."""
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    cols = detect_variant(ws)

    # Unmerge legend
    legend_row = find_legend_row(ws, cols)
    if legend_row:
        serial_col_letter = get_column_letter(cols['#'])
        for mr_str in [str(m) for m in list(ws.merged_cells.ranges)]:
            if str(legend_row) in mr_str and serial_col_letter in mr_str:
                try:
                    ws.unmerge_cells(mr_str)
                except Exception:
                    pass

    # Continue numbering
    last_data_row = find_last_data_row(ws, cols)
    last_serial = ws.cell(row=last_data_row, column=cols['#']).value
    if isinstance(last_serial, (int, float)):
        next_serial = int(last_serial) + 1
    else:
        next_serial = 1

    insert_at = (legend_row if legend_row else last_data_row + 1)
    n = len(rows)
    ws.insert_rows(insert_at, amount=n)

    for i, row in enumerate(rows):
        r = insert_at + i

        # Generous row height for wrapped text + ensure row is visible
        ws.row_dimensions[r].height = 90
        ws.row_dimensions[r].hidden = False

        # Populate values
        ws.cell(row=r, column=cols['#']).value = next_serial + i
        ws.cell(row=r, column=cols['name']).value = row.get('name', '')
        ws.cell(row=r, column=cols['indian']).value = row.get('indian', '')

        # Website (col E) — set hyperlink if URL provided
        website_url = (row.get('website') or '').strip()
        ws.cell(row=r, column=cols['website']).value = website_url
        if website_url and website_url.startswith(('http://', 'https://')):
            ws.cell(row=r, column=cols['website']).hyperlink = website_url

        # LinkedIn (col F) — single URL only, set as hyperlink
        # Accepts: full URL string OR empty/None for blank cell
        # If multiple URLs are passed (semicolon-separated), uses the FIRST one only
        linkedin_raw = (row.get('linkedin') or '').strip()
        if linkedin_raw:
            # Take first URL if multiple were passed
            first_url = linkedin_raw.split(';')[0].strip()
            # Strip any "Founders:" or "Founder:" prefix
            for prefix in ['Founders:', 'Founder:', 'Company:']:
                if first_url.startswith(prefix):
                    first_url = first_url[len(prefix):].strip()
            # Ensure URL has https:// prefix
            if first_url and not first_url.startswith(('http://', 'https://')):
                if first_url.startswith('linkedin.com') or first_url.startswith('in.linkedin.com'):
                    first_url = 'https://' + first_url
            # Only write if it looks like a real URL
            if first_url and ('linkedin.com' in first_url or first_url.startswith(('http://', 'https://'))):
                ws.cell(row=r, column=cols['linkedin']).value = first_url
                ws.cell(row=r, column=cols['linkedin']).hyperlink = first_url
            else:
                # Not a URL — leave blank rather than write text that won't be clickable
                ws.cell(row=r, column=cols['linkedin']).value = ''
        else:
            ws.cell(row=r, column=cols['linkedin']).value = ''

        ws.cell(row=r, column=cols['founded']).value = row.get('founded', '')
        ws.cell(row=r, column=cols['location']).value = row.get('location', '')
        ws.cell(row=r, column=cols['funding']).value = row.get('funding', '')
        ws.cell(row=r, column=cols['founders']).value = row.get('founders', '')
        ws.cell(row=r, column=cols['team_size']).value = row.get('team_size', '')
        ws.cell(row=r, column=cols['capabilities']).value = row.get('capabilities', '')
        ws.cell(row=r, column=cols['BALCO']).value = row.get('balco', '')
        ws.cell(row=r, column=cols['VAL-L']).value = row.get('val_l', '')
        ws.cell(row=r, column=cols['VAL-J']).value = row.get('val_j', '')
        ws.cell(row=r, column=cols['ESL']).value = row.get('esl', '')
        ws.cell(row=r, column=cols['HZL']).value = row.get('hzl', '')
        ws.cell(row=r, column=cols['Cairn']).value = row.get('cairn', '')
        ws.cell(row=r, column=cols['IOB']).value = row.get('iob', '')
        ws.cell(row=r, column=cols['verdict']).value = row.get('verdict', '')

        # Apply per-column formatting
        format_data_row(ws, r, cols)

        # Apply emoji-driven fills (overrides whatever was set in format_data_row for M-T)
        for col_key in ['BALCO', 'VAL-L', 'VAL-J', 'ESL', 'HZL', 'Cairn', 'IOB', 'verdict']:
            apply_emoji_fill(ws.cell(row=r, column=cols[col_key]))

    # Re-merge legend at new position
    if legend_row:
        new_legend_row = legend_row + n
        verdict_col_letter = get_column_letter(cols['verdict'])
        serial_col_letter = get_column_letter(cols['#'])
        ws.merge_cells(f"{serial_col_letter}{new_legend_row}:{verdict_col_letter}{new_legend_row}")
        # Ensure legend row is visible
        ws.row_dimensions[new_legend_row].hidden = False

    wb.save(workbook_path)
    return next_serial, next_serial + n - 1

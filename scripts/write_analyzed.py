"""
write_analyzed.py — Append rows into the Analyzed Startups sheet of
Ventures_Tracker.xlsx.

Format (per existing sheet):
  A: (gutter)
  B: Priority Tier
  C: Company
  D: Vedanta BU
  E: (gutter)
  F: Recommended Action
  G: Engagement Owner
  H: Rationale / Remarks
  I: Progress (left blank)

NO round-header rows. Append data rows directly.

Usage:
    from write_analyzed import append_to_analyzed

    rows = [
        {
            "tier": "P1 – Immediate",
            "company": "Xterra Robotics",
            "bu": "HZL / IOB",
            "action": "Pilot at Rampura Agucha for confined-space inspection. Validate stair-climbing capability before scaling.",
            "owner": "HZL Mines",
            "rationale": "Indigenous quadruped, IIT-K incubated. SVAN M2 + Cobot C1 directly applicable.",
        },
        ...
    ]
    append_to_analyzed("/path/to/workbook.xlsx", rows)
"""

from openpyxl import load_workbook
from copy import copy


CANONICAL_TIERS = {
    'P1 – Immediate',
    'P1 – Immediate (Strategic)',
    'P2 – BU-specific',
    'P2 – Conditional BU Evaluation',
    'P2 – Platform / Enabler (Conditional)',
    'P3 – Conditional / Procurement-led',
    'Deprioritise (CSR / Central only)',
    'Deprioritise',
}


def normalize_tier(tier):
    """Map any tier variant to canonical form."""
    t = (tier or '').strip()
    # Common typo / spacing fixes
    t = t.replace('Deprioritize', 'Deprioritise')
    t = t.replace('P1-Immediate', 'P1 – Immediate')
    t = t.replace('P1 – Immediate ', 'P1 – Immediate')  # trailing space
    if t == 'P2':
        t = 'P2 – BU-specific'
    if t.startswith('Deprioritise – Park'):
        t = 'Deprioritise'
    if t.startswith('Deprioritise (CSR only)') or t.startswith('Deprioritise (Central Legal only)'):
        t = 'Deprioritise (CSR / Central only)'
    if t in CANONICAL_TIERS:
        return t
    # Fallback — accept as-is if unmappable
    return tier


def find_last_data_row(ws):
    """Find the last non-empty row in the Analyzed Startups sheet."""
    last = 2  # header is at row 2
    for r in range(3, ws.max_row + 2):
        company = ws.cell(row=r, column=3).value  # column C
        if company and isinstance(company, str) and company.strip():
            last = r
    return last


def find_reference_data_row(ws):
    """Find a populated data row to copy formatting from."""
    for r in range(3, ws.max_row + 1):
        company = ws.cell(row=r, column=3).value
        if company and isinstance(company, str) and company.strip() and 'company' not in company.lower():
            return r
    return 3  # fallback


def append_to_analyzed(workbook_path, rows, sheet_name='Analyzed Startups'):
    """Append a list of row-dicts to Analyzed Startups sheet.

    Each row dict expects: tier, company, bu, action, owner, rationale.
    """
    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        # Create the sheet with headers if missing
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=2, column=2).value = 'Priority Tier'
        ws.cell(row=2, column=3).value = 'Company'
        ws.cell(row=2, column=4).value = 'Vedanta BU'
        ws.cell(row=2, column=6).value = 'Recommended Action'
        ws.cell(row=2, column=7).value = 'Engagement Owner'
        ws.cell(row=2, column=8).value = 'Rationale / Remarks'
        ws.cell(row=2, column=9).value = 'Progress'
    else:
        ws = wb[sheet_name]

    last_row = find_last_data_row(ws)
    ref_row = find_reference_data_row(ws)
    insert_at = last_row + 1

    for i, row in enumerate(rows):
        r = insert_at + i

        # Ensure row is visible
        ws.row_dimensions[r].hidden = False

        # Copy formatting from reference data row
        for col in range(1, 10):
            src = ws.cell(row=ref_row, column=col)
            dst = ws.cell(row=r, column=col)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)

        # Populate
        ws.cell(row=r, column=2).value = normalize_tier(row.get('tier', ''))
        ws.cell(row=r, column=3).value = row.get('company', '')
        ws.cell(row=r, column=4).value = row.get('bu', '–')
        ws.cell(row=r, column=6).value = row.get('action', '')
        ws.cell(row=r, column=7).value = row.get('owner', '–')
        ws.cell(row=r, column=8).value = row.get('rationale', '')
        # Column I (Progress) left blank intentionally

    wb.save(workbook_path)
    return insert_at, insert_at + len(rows) - 1


# Word count enforcement helper
def enforce_caps(row):
    """Check Action and Rationale don't exceed reasonable lengths."""
    warnings = []
    caps = {'action': 35, 'rationale': 40}
    for key, cap in caps.items():
        v = row.get(key, '')
        if v and isinstance(v, str):
            wc = len(v.split())
            if wc > cap:
                warnings.append(f"{key}: {wc} words (cap {cap})")
    return warnings
